import openpyxl


def create_xl(date_time):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    wb.create_sheet("Счет")
    wb.create_sheet("Итог")

    # Лист Счет
    ws = wb.worksheets[0]
    
    ws.cell(row=1, column=1, value="Пользователь")
    ws.cell(row=1, column=2, value="Наименование")
    ws.cell(row=1, column=3, value="Количество в заказе")
    ws.cell(row=1, column=4, value="Цена книги")
    ws.cell(row=1, column=5, value="Общая цена")
    ws.cell(row=1, column=6, value="Итоговая сумма")
    
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 14

    # Лист Итог
    ws = wb.worksheets[1]

    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="Всего")
    ws.cell(row=1, column=3, value=0)
    
    ws.cell(row=2, column=1, value="Номер")
    ws.cell(row=2, column=2, value="Пользователь")
    ws.cell(row=2, column=3, value="Сумма")
    
    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10

    wb.save(f'Data/{str(date_time)}.xlsx')


def edit_xl(date_time, user_name, books, prices, countes, itog):
    try:
        wb = openpyxl.load_workbook(f"Data/{date_time}.xlsx")
    except:
        create_xl(date_time)
        wb = openpyxl.load_workbook(f"Data/{date_time}.xlsx")

    ws = wb.worksheets[0]
    ws.append([''])
    ws.append([f"{user_name}", books[0], int(countes[0]), int(prices[0]),
               int(prices[0]) * int(countes[0]), int(itog)])
    for i in range(1, len(books)):
        ws.append(['', f"{books[i]}", int(countes[i]), int(prices[i]), int(prices[i]) * int(countes[i])])
    wb.save(f'Data/{date_time}.xlsx')


def edit_total(date_time, full_name, amount):
    try:
        wb = openpyxl.load_workbook(f"Data/{date_time}.xlsx")
    except:
        create_xl(date_time)
        wb = openpyxl.load_workbook(f"Data/{date_time}.xlsx")

    ws = wb.worksheets[1]
    max_rows = ws.max_row
    total_amount = ws.cell(row=1, column=3).value + int(amount)
    ws.cell(row=1, column=3, value=total_amount)
    err = 1
    for i_row in range(1, max_rows+1):
        if ws.cell(row=i_row, column=2).value == full_name:
            ammo = ws.cell(row=i_row, column=3).value
            ws.cell(row=i_row, column=3, value=ammo+int(amount))
            err = 0
            break
    if err == 1:
        ws.append([max_rows-1, full_name, int(amount)])
    wb.save(f'Data/{str(date_time)}.xlsx')
